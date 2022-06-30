import io
import json
import math
import os
import random
import re
import shutil
import subprocess
import uuid
from datetime import datetime
from time import time

import pytz
from flask import render_template, flash, current_app, send_from_directory, request, redirect, url_for, send_file
from flask_login import login_required, current_user
from openpyxl import load_workbook
from sqlalchemy.orm import load_only
from werkzeug.utils import secure_filename
from xlrd import open_workbook

from common.logger import log
from config import Config
from qti.itemservice.itemservice import ItemService
from qti.loader.itemloader import ItemLoader
from . import item
from .forms import ItemSearchForm, ItemLoadForm, ItemListForm, FileLoadForm, ItemEditForm, \
    ItemEditSubForm, ItemEditExplanationForm, ItemAssessmentSearchForm
from .. import db
from ..api.response import success
from ..decorators import permission_required, permission_required_or_multiple
from ..models import Codebook, Item, Permission, Choices, ItemExplanation, Testset, TestletHasItem
from app.api.jwplayer import get_signed_player, jwt_signed_url
from app.api.errors import bad_request
import re

@item.route('/<int:id>', methods=['GET'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def info(id):
    item = Item.query.filter_by(id=id).first()
    if item:
        return render_template("item/item_info.html", item=item)


@item.route('/<int:item_id>/peek', methods=['GET'])
@login_required
# @permission_required(Permission.ITEM_EXEC)
@permission_required_or_multiple(Permission.ITEM_EXEC, Permission.ASSESSMENT_READ)
def peek(item_id):
    # TODO - Needs to procetect this with hashing
    response = {
        'result': 'success',
    }
    rendered_item = ''
    qti_item_obj = Item.query.filter_by(id=item_id).first()

    processed = None
    correct_r_value = None
    if os.environ.get("DEBUG_RENDERING", 'false') == 'false':
        try:
            item_service = ItemService(qti_item_obj.file_link)
            qti_item = item_service.get_item()
            rendered_item = qti_item.to_html()
            response['type'] = qti_item.get_interaction_type()
            response['cardinality'] = qti_item.get_cardinality()
            response['object_variables'] = qti_item.get_interaction_object_variables()

            response = {"RESPONSE": {"base": {"identifier": ""}}}
            qti_xml = item_service.get_qti_xml_path()
            processing_php = current_app.config['QTI_RSP_PROCESSING_PHP']
            parameter = json.dumps({'response': response, 'qtiFilename': qti_xml})

            result = subprocess.run(['php', processing_php, parameter], stdout=subprocess.PIPE)
            processed = result.stdout.decode("utf-8")
            processed = json.loads(processed)
            correct_r_value = parse_correct_response(processed.get('correctResponses'))

        except Exception as e:
            print(e)
    else:
        item_service = ItemService(qti_item_obj.file_link)
        qti_item = item_service.get_item()
        rendered_item = qti_item.to_html()
        response['type'] = qti_item.get_interaction_type()
        response['cardinality'] = qti_item.get_cardinality()
        response['object_variables'] = qti_item.get_interaction_object_variables()

        response = {"RESPONSE": {"base": {"identifier": ""}}}
        qti_xml = item_service.get_qti_xml_path()
        processing_php = current_app.config['QTI_RSP_PROCESSING_PHP']
        parameter = json.dumps({'response': response, 'qtiFilename': qti_xml})

        result = subprocess.run(['php', processing_php, parameter], stdout=subprocess.PIPE)
        processed = result.stdout.decode("utf-8")
        processed = json.loads(processed)
        correct_r_value = parse_correct_response(processed.get('correctResponses'))

    qti_item_obj.correct_answer = correct_r_value
    rendered_template = render_template("item/item_peek.html", item=qti_item_obj, hide_answer=1)
    if rendered_item:
        rendered_template = rendered_template.replace('Preview not available', rendered_item)

    response['html'] = rendered_template
    return success(response)


@item.route('/import', methods=['GET'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def item_import():
    grade = request.args.get('grade')
    subject = request.args.get('subject')
    file_name = request.args.get('file_name')
    items = request.args.get('items')
    error = request.args.get('error')

    default_choices = Choices()

    file_form = FileLoadForm()
    file_form.grade.choices = default_choices.grade_choices
    file_form.subject.choices = default_choices.subject_choices

    if error:
        flash(error)

    if file_name:
        file_form.grade.default = grade
        file_form.subject.default = subject
        file_form.process()
    else:
        pass

    if items:
        item_form = ItemLoadForm()
        item_form.level.choices = default_choices.level_choices
        item_form.category.choices = set_choices_child(int(subject))
        item_form.subcategory.choices = [(0, ' ')]

        item_list = items.split(',')
        item_db = Item.query.filter(Item.id.in_(item_list)).all()
        item_form = populate_item_form(item_form, item_db)
        pass
    else:
        item_form = None

    return render_template('item/item_import.html', file_form=file_form, item_form=item_form, file_name=file_name)


@item.route('/import/load', methods=['POST'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def item_import_load():
    my_files, uploaded_files, rejected_files, items, listform = [], [], [], [], []
    default_choices = Choices()

    file_form = FileLoadForm()
    file_form.grade.choices = default_choices.grade_choices
    file_form.subject.choices = default_choices.subject_choices

    if file_form.validate_on_submit():
        my_files = request.files.getlist("items_file")
        parsed_items = []
        for f in my_files:
            f_file_name = f.filename
            if f and allowed_file(f.filename):
                my_guid = str(uuid.uuid4())
                file_name = my_guid + secure_filename(f.filename)
                item_file = os.path.join(current_app.config['UPLOAD_FOLDER'], file_name)
                f.save(item_file)
                uploaded_files.append(f)
                extension = file_name.rsplit('.', 1)[1].lower()
                if extension == 'xml':
                    parsed = parse_qti_xml(item_file, my_guid, file_name)
                    if parsed is not None:
                        parsed_items.append(parsed)
                elif extension == 'zip':
                    parsed_items = parsed_items + parse_qti_package_zip(item_file, file_name)
                elif extension in ('xls', 'xlsx'):
                    package_file = get_package_from_excel(item_file)
                    parsed_items = parsed_items + parse_qti_package_zip(package_file, os.path.basename(package_file))

                for parsed in parsed_items:
                    a_item = Item(GUID=parsed.get('guid'),
                                  grade=file_form.grade.data,
                                  subject=file_form.subject.data,
                                  imported_by=current_user.id)
                    qti_item = parsed.get('qti_item')
                    a_item.set_value_from_qti(qti_item)
                    a_item.correct_r_value = parsed.get('correct_r_value')
                    a_item.outcome_score = parsed.get('outcome_score')
                    db.session.add(a_item)
                    items.append(a_item)
                log.info("New Item File {} imported.".format(file_name))
            else:
                rejected_files.append(f)
                log.info("Reject {} file importing".format(f.filename))
                return redirect(url_for('item.item_import', error=rejected_files))
            db.session.commit()
            item_list_string = [str(i.id) for i in items]
            item_list_string = ','.join(item_list_string)
            return redirect(url_for('item.item_import',
                                    grade=file_form.grade.data,
                                    subject=file_form.subject.data,
                                    file_name=f_file_name,
                                    items=item_list_string))
    return redirect(url_for('item.item_import', error="File load validation error"))


@item.route('/import/import', methods=['POST'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def item_import_import():
    default_choices = Choices()

    item_form = ItemLoadForm()
    item_form.level.choices = default_choices.level_choices
    item_form.category.choices = default_choices.category_choices
    item_form.subcategory.choices = default_choices.subcategory_choices

    # Initialise choices for FieldList forms for validation
    if item_form.items is not None:
        for entry in item_form.items.entries:
            entry.form.level.choices = default_choices.level_choices
            entry.form.category.choices = default_choices.category_choices
            entry.form.subcategory.choices = default_choices.subcategory_choices

    if item_form.validate_on_submit():
        if item_form.items is not None:
            for data in item_form.items.data:
                a_item = Item.query.filter_by(id=int(data.get('item_id'))).first()
                a_item.name = data.get('item_name')
                a_item.correct_answer = data.get('correct_answer')
                a_item.level = data.get('level')
                a_item.category = data.get('category')
                a_item.subcategory = data.get('subcategory')
                if a_item.level != '0':  # check if level data is blank, switch active when data exisiting
                    a_item.active = True
                a_item.modified_by = current_user.id
                a_item.modified_time = datetime.now(pytz.utc)
                db.session.commit()
            flash("Your items are imported successfully.")
        else:
            flash("No items are imported. Please check your package file imported.")
        return redirect(url_for('item.item_import'))
    return redirect(url_for('item.item_import', error="File load validation error"))


def populate_item_form(form, items=None):
    default_choices = Choices()

    level_choices = default_choices.level_choices
    if items:
        category_choices = set_choices_child(items[0].subject)
    else:
        category_choices = [(0, ' ')]
    subcategory_choices = [(0, ' ')]

    if items is not None:
        while len(form.items) > 0:
            form.items.pop_entry()

        for item in items:
            item_form = ItemListForm(level_choices=level_choices,
                                     category_choices=category_choices,
                                     subcategory_choices=subcategory_choices)
            item_form.item_id = item.id
            item_form.item_name = item.name
            item_form.correct_answer = item.correct_answer
            item_form.grade = Codebook.get_code_name(item.grade)
            item_form.subject = Codebook.get_code_name(item.subject)
            form.items.append_entry(item_form)

        for item_form in form.items:
            item_form.form.level.choices = level_choices
            item_form.form.category.choices = category_choices
            item_form.form.subcategory.choices = subcategory_choices

    return form


@item.route('/list', methods=['GET', 'POST'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def item_list():
    items = []
    search_form = ItemSearchForm()
    default_choices = Choices()
    search_form.grade.choices = default_choices.grade_choices
    search_form.subject.choices = default_choices.subject_choices
    search_form.level.choices = default_choices.level_choices
    if search_form.subject.data:
        search_form.category.choices = set_choices_child(search_form.subject.data)
    else:
        search_form.category.choices = [(0, ' ')]
    if search_form.category.data:
        search_form.subcategory.choices = set_choices_child(search_form.category.data)
    else:
        search_form.subcategory.choices = [(0, ' ')]

    query = Item.query
    if search_form.validate_on_submit():
        if search_form.grade.data:
            query = query.filter_by(grade=search_form.grade.data)
        if search_form.subject.data:
            query = query.filter_by(subject=search_form.subject.data)
        if search_form.level.data:
            query = query.filter_by(level=search_form.level.data)
        if search_form.category.data:
            query = query.filter_by(category=search_form.category.data)
        if search_form.subcategory.data:
            query = query.filter_by(subcategory=search_form.subcategory.data)
        if search_form.byme.data:
            query = query.filter_by(modified_by=current_user.id)
        if search_form.active.data:
            query = query.filter_by(active=search_form.active.data)
        items = query.order_by(Item.id.desc()).all()
        flash('Found {} item(s)'.format(len(items)))
    return render_template('item/item_list.html', form=search_form, items=items)


@item.route('/assessment/list', methods=['GET'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def item_assessment_list():
    items = []
    search_form = ItemAssessmentSearchForm()
    search_form.test_type.data = Codebook.get_code_id('Naplan')
    # default setting value into test_center list
    branch_id = current_user.get_branch_id()
    if branch_id and current_user.username != 'All':
        search_form.test_center.data = branch_id
    else:
        search_form.test_center.data = None
    search_form.year.data = None

    query = Item.query
    if search_form.validate_on_submit():
        if search_form.grade.data:
            query = query.filter_by(grade=search_form.grade.data)
        if search_form.subject.data:
            query = query.filter_by(subject=search_form.subject.data)
        if search_form.level.data:
            query = query.filter_by(level=search_form.level.data)
        if search_form.category.data:
            query = query.filter_by(category=search_form.category.data)
        if search_form.subcategory.data:
            query = query.filter_by(subcategory=search_form.subcategory.data)
        if search_form.byme.data:
            query = query.filter_by(modified_by=current_user.id)
        if search_form.active.data:
            query = query.filter_by(active=search_form.active.data)
        items = query.order_by(Item.id.desc()).all()
        flash('Found {} item(s)'.format(len(items)))
    return render_template('item/item_assessment_list.html', form=search_form, items=items)


@item.route('/assessment/answer/update', methods=['GET'])
@login_required
@permission_required(Permission.ADMIN)
def item_assessment_update_answer():
    search_form = ItemAssessmentSearchForm()
    search_form.test_type.data = Codebook.get_code_id('Naplan')
    # default setting value into test_center list
    branch_id = current_user.get_branch_id()
    if branch_id and current_user.username != 'All':
        search_form.test_center.data = branch_id
    else:
        search_form.test_center.data = None
    search_form.year.data = None

    return render_template('item/item_assessment_answer.html', form=search_form)


@item.route('/assessment/list', methods=['POST'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def item_assessment_search_list():
    result = []
    testset_id = request.form.get('testset_id')
    if testset_id is None:
        return bad_request()

    testset = Testset.query.filter_by(id=testset_id).first()
    if testset is None:
        return bad_request()

    branching = json.dumps(testset.branching)
    ends = [m.end() for m in re.finditer('"id":', branching)]
    for end in ends:
        comma = branching.find(',', end)
        testlet_id = int(branching[end:comma])

        items = Item.query.join(TestletHasItem, Item.id == TestletHasItem.item_id) \
            .filter(TestletHasItem.testlet_id == testlet_id) \
            .order_by(TestletHasItem.order).all()


        for item in items:
            i = {'id':item.id,
                 'name':item.name,
                 'correct_answer':item.correct_answer,
                 'grade':Codebook.get_code_name(item.grade),
                 'subject':Codebook.get_code_name(item.subject),
                 'level':Codebook.get_code_name(item.level),
                 'category':Codebook.get_code_name(item.category),
                 'subcategory':Codebook.get_code_name(item.subcategory),
                 'active':item.active
                 }
            result.append(i)

    return success(result)


@item.route('/edit/edit', methods=['POST'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def item_edit_edit():
    default_choices = Choices()

    item_form = ItemEditForm()
    item_form.grade.choices = default_choices.grade_choices
    item_form.subject.choices = default_choices.subject_choices
    item_form.level.choices = default_choices.level_choices
    item_form.category.choices = default_choices.category_choices
    item_form.subcategory.choices = default_choices.subcategory_choices

    # Initialise choices for FieldList forms for validation
    if item_form.items is not None:
        for entry in item_form.items.entries:
            entry.form.grade.choices = default_choices.grade_choices
            entry.form.subject.choices = default_choices.subject_choices
            entry.form.level.choices = default_choices.level_choices
            entry.form.category.choices = default_choices.category_choices
            entry.form.subcategory.choices = default_choices.subcategory_choices

    if item_form.validate_on_submit():
        if item_form.items is not None:
            for data in item_form.items.data:
                a_item = Item.query.filter_by(id=int(data.get('item_id'))).first()
                a_item.name = data.get('item_name')
                a_item.correct_answer = data.get('correct_answer')
                a_item.grade = data.get('grade')
                a_item.subject = data.get('subject')
                a_item.level = data.get('level')
                a_item.category = data.get('category')
                a_item.subcategory = data.get('subcategory')
                a_item.active = data.get('active')
                a_item.modified_by = current_user.id
                a_item.modified_time = datetime.now(pytz.utc)
                db.session.add(a_item)
            db.session.commit()
            flash("Your items are modified successfully.")
        else:
            flash("No items are modified. Please search again and try to edit.")

        return redirect(url_for('item.manage'))
    return redirect(url_for('item.manage', error="Item Edit - Form validation error"))


@item.route('/manage', methods=['GET'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def manage():
    grade = request.args.get('grade')
    subject = request.args.get('subject')
    level = request.args.get('level')
    category = request.args.get('category')
    subcategory = request.args.get('subcategory')
    byme = request.args.get('byme')
    active = request.args.get('active')
    submit = request.args.get('submit')

    error = request.args.get('error')

    default_choices = Choices()
    search_form = ItemSearchForm()
    search_form.grade.choices = default_choices.grade_choices
    search_form.subject.choices = default_choices.subject_choices
    search_form.level.choices = default_choices.level_choices
    if subject:
        search_form.category.choices = set_choices_child(subject)
    else:
        search_form.category.choices = [(0, ' ')]
    if category:
        search_form.subcategory.choices = set_choices_child(category)
    else:
        search_form.subcategory.choices = [(0, ' ')]
    if error:
        flash(error)
    search_form.grade.default = grade
    search_form.subject.default = subject
    search_form.level.default = level
    search_form.category.default = category
    search_form.subcategory.default = subcategory
    search_form.process()
    item_form = None
    if submit:
        query = Item.query.options(load_only("id"))
        if grade != '0':
            query = query.filter_by(grade=grade)
        if subject != '0':
            query = query.filter_by(subject=subject)
        if level != '0':
            query = query.filter_by(level=level)
        if category != '0':
            query = query.filter_by(category=category)
        if subcategory != '0':
            query = query.filter_by(subcategory=subcategory)
        if byme == 'y':
            query = query.filter_by(imported_by=current_user.id)
        if active == 'y':
            query = query.filter_by(active=active)
        items = query.order_by(Item.modified_time.desc()).all()
        if items:
            flash('Found {} item(s)'.format(len(items)))
            item_form = ItemEditForm()
            item_form.grade.choices = default_choices.grade_choices
            item_form.subject.choices = default_choices.subject_choices
            item_form.level.choices = default_choices.level_choices

            if subject:
                item_form.category.choices = set_choices_child(subject)
            else:
                item_form.category.choices = [(0, ' ')]
            if category:
                item_form.subcategory.choices = set_choices_child(category)
            else:
                item_form.subcategory.choices = [(0, ' ')]

            item_form.grade.default = grade
            item_form.subject.default = subject
            item_form.level.default = level
            item_form.category.default = category
            item_form.subcategory.default = subcategory
            item_form.process()
            #
            # item_db = Item.query.filter(Item.id.in_(item_list)).all()
            # item_form = populate_item_edit_form(item_form, item_db)
            item_form = populate_item_edit_form(item_form, items)
        else:
            item_form = None
    return render_template('item/manage.html', item_form=item_form, search_form=search_form)


@item.route('/export', methods=['GET'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def item_export():
    from shutil import copyfile

    query = Item.query
    grade = request.args.get('grade', '0')
    subject = request.args.get('subject', '0')
    level = request.args.get('level', '0')
    category = request.args.get('category', '0')
    subcategory = request.args.get('subcategory', '0')
    byme = request.args.get('byme', '0')
    active = request.args.get('active', '0')
    if grade != '0':
        query = query.filter_by(grade=grade)
    if subject != '0':
        query = query.filter_by(subject=subject)
    if level != '0':
        query = query.filter_by(level=level)
    if category != '0':
        query = query.filter_by(category=category)
    if subcategory != '0':
        query = query.filter_by(subcategory=subcategory)
    if byme != '0':
        query = query.filter_by(modified_by=current_user.id)
    if active != '0':
        query = query.filter_by(active=active)
    items = query.order_by(Item.id.desc()).all()
    resources = []
    manifest_id = get_identifier()
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    manifest_dir = os.path.join(Config.IMPORT_TEMP_DIR, timestamp, manifest_id)
    os.makedirs(manifest_dir)
    duplication_counter = 0
    for a_item in items:
        file_dir = os.path.join(current_app.config['STORAGE_DIR'], a_item.file_link)
        file_dir = os.path.join(file_dir, 'itemContent/en-US')
        resource_dir = os.path.join(manifest_dir, a_item.file_link)
        # Handle duplicated ID
        if os.path.exists(resource_dir):
            duplication_counter += 1
            resource_dir = "%s_dupilacted_%d" % (resource_dir, duplication_counter)
        os.mkdir(resource_dir)
        a_resource = {'identifier': a_item.file_link, 'files': []}
        for a_file in os.listdir(file_dir):
            file_path = os.path.join(file_dir, a_file)
            if os.path.isfile(file_path):
                copyfile(file_path, os.path.join(resource_dir, a_file))
                a_resource['files'].append(a_file)
        resources.append(a_resource)
    log.warn("%d duplicated items found" % duplication_counter)
    manifest_xml = render_template('import/imsmanifest_export.xml', identifier=manifest_id, resources=resources)

    with open(os.path.join(manifest_dir, 'imsmanifest.xml'), 'w+', encoding='utf-8') as f:
        f.write(manifest_xml)

    zip_path = os.path.normpath('%s/%s.zip' % (Config.IMPORT_TEMP_DIR, manifest_id))
    shutil.make_archive(zip_path.replace('.zip', ''), 'zip', manifest_dir)
    with open(zip_path, 'rb') as bites:
        rsp = send_file(
            io.BytesIO(bites.read()), as_attachment=True,
            attachment_filename='%s.zip' % manifest_id,
            mimetype='application/zip'
        )
    return rsp

@item.route('/assessment/export', methods=['GET'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def item_assessment_export():
    from shutil import copyfile

    items = []
    testset_id = request.args.get('testset_id', '0')
    testset = Testset.query.filter_by(id=testset_id).first()
    branching = json.dumps(testset.branching)
    ends = [m.end() for m in re.finditer('"id":', branching)]
    for end in ends:
        comma = branching.find(',', end)
        testlet_id = int(branching[end:comma])

        item = Item.query.join(TestletHasItem, Item.id == TestletHasItem.item_id) \
            .filter(TestletHasItem.testlet_id == testlet_id) \
            .order_by(TestletHasItem.order).all()
        items.extend(item)

    resources = []
    manifest_id = get_identifier()
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    manifest_dir = os.path.join(Config.IMPORT_TEMP_DIR, timestamp, manifest_id)
    os.makedirs(manifest_dir)
    duplication_counter = 0
    for a_item in items:
        file_dir = os.path.join(current_app.config['STORAGE_DIR'], a_item.file_link)
        file_dir = os.path.join(file_dir, 'itemContent/en-US')
        resource_dir = os.path.join(manifest_dir, a_item.file_link)
        # Handle duplicated ID
        if os.path.exists(resource_dir):
            duplication_counter += 1
            resource_dir = "%s_dupilacted_%d" % (resource_dir, duplication_counter)
        os.mkdir(resource_dir)
        a_resource = {'identifier': a_item.file_link, 'files': []}
        for a_file in os.listdir(file_dir):
            file_path = os.path.join(file_dir, a_file)
            if os.path.isfile(file_path):
                copyfile(file_path, os.path.join(resource_dir, a_file))
                a_resource['files'].append(a_file)
        resources.append(a_resource)
    log.warn("%d duplicated items found" % duplication_counter)
    manifest_xml = render_template('import/imsmanifest_export.xml', identifier=manifest_id, resources=resources)

    with open(os.path.join(manifest_dir, 'imsmanifest.xml'), 'w+', encoding='utf-8') as f:
        f.write(manifest_xml)

    zip_path = os.path.normpath('%s/%s.zip' % (Config.IMPORT_TEMP_DIR, manifest_id))
    shutil.make_archive(zip_path.replace('.zip', ''), 'zip', manifest_dir)
    with open(zip_path, 'rb') as bites:
        rsp = send_file(
            io.BytesIO(bites.read()), as_attachment=True,
            attachment_filename='%s.zip' % manifest_id,
            mimetype='application/zip'
        )
    return rsp


def populate_item_edit_form(form, items=None):
    default_choices = Choices()

    grade_choices = default_choices.grade_choices
    subject_choices = default_choices.subject_choices
    level_choices = default_choices.level_choices
    if items:
        category_choices = set_choices_child(items[0].subject)
    else:
        category_choices = [(0, ' ')]
    subcategory_choices = [(0, ' ')]

    if items is not None:

        while len(form.items) > 0:
            form.items.pop_entry()

        item_grade, item_subject, item_level, item_category, item_subcategory = [], [], [], [], []
        for item in items:
            item_sub_form = ItemEditSubForm(grade_choices=grade_choices,
                                            subject_choices=subject_choices,
                                            level_choices=level_choices,
                                            category_choices=category_choices,
                                            subcategory_choices=subcategory_choices)
            form.items.append_entry(item_sub_form)

            if item.grade is not None:
                item_grade.append(item.grade)
            else:
                item_grade.append('0')

            if item.subject is not None:
                item_subject.append(item.subject)
            else:
                item_subject.append('0')

            if item.level is not None:
                item_level.append(item.level)
            else:
                item_level.append('0')

            if item.category is not None:
                item_category.append(item.category)
            else:
                item_category.append('0')

            if item.subcategory is not None:
                item_subcategory.append(item.subcategory)
            else:
                item_subcategory.append('0')

        i = 0
        for sub_form in form.items:
            sub_form.form.grade.choices = grade_choices
            sub_form.form.subject.choices = subject_choices
            sub_form.form.level.choices = level_choices
            if item_subject[i]:
                sub_form.form.category.choices = set_choices_child(item_subject[i])
            else:
                sub_form.form.category.choices = [(0, ' ')]
            if item_category[i]:
                sub_form.form.subcategory.choices = set_choices_child(item_category[i])
            else:
                sub_form.form.subcategory.choices = [(0, ' ')]

            sub_form.form.item_id.default = items[i].id
            sub_form.form.item_name.default = items[i].name
            sub_form.form.correct_answer.default = items[i].correct_answer
            sub_form.form.active.default = items[i].active

            sub_form.form.grade.default = item_grade[i]
            sub_form.form.subject.default = item_subject[i]
            sub_form.form.level.default = item_level[i]
            sub_form.form.category.default = item_category[i]
            sub_form.form.subcategory.default = item_subcategory[i]
            sub_form.form.process()
            i = i + 1

    return form


@item.route('/<int:item_id>/preview', methods=['GET'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def preview(item_id):
    rendered_preview = None
    from qti.itemservice.itemservice import ItemService
    qti_item_obj = Item.query.filter_by(id=item_id).first()
    if os.environ.get("DEBUG_RENDERING", 'false') == 'false':
        try:
            item_service = ItemService(qti_item_obj.file_link)
            qti_item = item_service.get_item()
            rendered_preview = qti_item.to_html()
        except Exception as e:
            print(e)
    else:
        item_service = ItemService(qti_item_obj.file_link)
        qti_item = item_service.get_item()
        rendered_preview = qti_item.to_html()
    rendered_template = render_template("item/item_info.html", item=qti_item_obj)
    if rendered_preview:
        rendered_template = rendered_template.replace('Preview not available', rendered_preview)
    return rendered_template


@item.route('/<int:item_id>/review/<int:test_type>', methods=['GET'])
@login_required
@permission_required_or_multiple(Permission.ITEM_EXEC, Permission.ASSESSMENT_READ)
def review(item_id, test_type=None):
    rendered_preview = None
    from qti.itemservice.itemservice import ItemService
    qti_item_obj = Item.query.filter_by(id=item_id).first()

    processed = None
    correct_r_value = None
    if os.environ.get("DEBUG_RENDERING", 'false') == 'false':
        try:
            item_service = ItemService(qti_item_obj.file_link)
            qti_item = item_service.get_item()
            rendered_preview = qti_item.to_html()

            response = {"RESPONSE": {"base": {"identifier": ""}}}
            qti_xml = item_service.get_qti_xml_path()
            processing_php = current_app.config['QTI_RSP_PROCESSING_PHP']
            parameter = json.dumps({'response': response, 'qtiFilename': qti_xml})

            result = subprocess.run(['php', processing_php, parameter], stdout=subprocess.PIPE)
            processed = result.stdout.decode("utf-8")
            processed = json.loads(processed)
            correct_r_value = parse_correct_response(processed.get('correctResponses'))
        except Exception as e:
            print(e)
    else:
        item_service = ItemService(qti_item_obj.file_link)
        qti_item = item_service.get_item()
        rendered_preview = qti_item.to_html()

        response = {"RESPONSE": {"base": {"identifier": ""}}}
        qti_xml = item_service.get_qti_xml_path()
        processing_php = current_app.config['QTI_RSP_PROCESSING_PHP']
        parameter = json.dumps({'response': response, 'qtiFilename': qti_xml})

        result = subprocess.run(['php', processing_php, parameter], stdout=subprocess.PIPE)
        processed = result.stdout.decode("utf-8")
        processed = json.loads(processed)
        correct_r_value = parse_correct_response(processed.get('correctResponses'))

    qti_item_obj.correct_answer = correct_r_value
    rendered_template = render_template("item/item_peek.html", item=qti_item_obj, test_type=test_type)

    if Codebook.get_code_name(test_type) != 'CBOCTT' and Codebook.get_code_name(test_type) != 'CBSTT' \
            and Codebook.get_code_name(test_type) != 'Naplan':
        rendered_preview = None
    if rendered_preview:
        rendered_template = rendered_template.replace('Preview not available', rendered_preview)
    return rendered_template


@item.route('/<int:item_id>/review/<int:test_type>', methods=['POST'])
@login_required
@permission_required_or_multiple(Permission.ITEM_EXEC, Permission.ASSESSMENT_READ)
def review_async(item_id, test_type=None):
    rendered_preview = None
    from qti.itemservice.itemservice import ItemService
    qti_item_obj = Item.query.filter_by(id=item_id).first()

    processed = None
    correct_r_value = None
    if os.environ.get("DEBUG_RENDERING", 'false') == 'false':
        try:
            item_service = ItemService(qti_item_obj.file_link)
            qti_item = item_service.get_item()
            rendered_preview = qti_item.to_html()

            response = {"RESPONSE": {"base": {"identifier": ""}}}
            qti_xml = item_service.get_qti_xml_path()
            processing_php = current_app.config['QTI_RSP_PROCESSING_PHP']
            parameter = json.dumps({'response': response, 'qtiFilename': qti_xml})

            result = subprocess.run(['php', processing_php, parameter], stdout=subprocess.PIPE)
            processed = result.stdout.decode("utf-8")
            processed = json.loads(processed)
            correct_r_value = parse_correct_response(processed.get('correctResponses'))
        except Exception as e:
            print(e)
    else:
        item_service = ItemService(qti_item_obj.file_link)
        qti_item = item_service.get_item()
        rendered_preview = qti_item.to_html()

        response = {"RESPONSE": {"base": {"identifier": ""}}}
        qti_xml = item_service.get_qti_xml_path()
        processing_php = current_app.config['QTI_RSP_PROCESSING_PHP']
        parameter = json.dumps({'response': response, 'qtiFilename': qti_xml})

        result = subprocess.run(['php', processing_php, parameter], stdout=subprocess.PIPE)
        processed = result.stdout.decode("utf-8")
        processed = json.loads(processed)
        correct_r_value = parse_correct_response(processed.get('correctResponses'))

    qti_item_obj.correct_answer = correct_r_value
    if Codebook.get_code_name(test_type) != 'CBOCTT' and Codebook.get_code_name(test_type) != 'CBSTT' \
            and Codebook.get_code_name(test_type) != 'Naplan':
        rendered_preview = None

    result = {'item': qti_item_obj, 'test_type': test_type, 'rendered_preview': rendered_preview}
    return result


@item.route('/<int:item_id>/rendered', methods=['GET'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def rendered(item_id):
    response = {
        'result': 'success',
    }
    rendered_item = ''
    qti_item_obj = Item.query.filter_by(id=item_id).first()
    if os.environ.get("DEBUG_RENDERING", 'false') == 'false':
        try:
            item_service = ItemService(qti_item_obj.file_link)
            qti_item = item_service.get_item()
            rendered_item = qti_item.to_html()
            response['type'] = qti_item.get_interaction_type()
            response['cardinality'] = qti_item.get_cardinality()
            response['object_variables'] = qti_item.get_interaction_object_variables()
        except Exception as e:
            print(e)
    else:
        item_service = ItemService(qti_item_obj.file_link)
        qti_item = item_service.get_item()
        rendered_item = qti_item.to_html()
        response['type'] = qti_item.get_interaction_type()
        response['cardinality'] = qti_item.get_cardinality()
        response['object_variables'] = qti_item.get_interaction_object_variables()
    rendered_template = render_template("item/item_info.html", item=qti_item_obj)
    if rendered_item:
        rendered_template = rendered_template.replace('Preview not available', rendered_item)

    response['html'] = rendered_template
    media_id_match = re.search(r"http://jwplayer-id/([a-zA-Z0-9]+)", rendered_template)
    if media_id_match:
        remained_sec = 60
        # Max time remained set to 50 minutes
        remained_sec = min(max(remained_sec, 0), 3000)
        # Link is valid for remained_sec but normalized to 5 minutes to promote better caching
        expires = math.ceil((time() + remained_sec) / 300) * 300
        media_id = media_id_match.group(1)
        jw_key = current_app.config['JWAPI_CREDENTIAL']
        player_id = current_app.config['JWPLAYER_ID']
        signed_player_url = get_signed_player(player_id, jw_key, expires)
        path = "/v2/media/{media_id}".format(media_id=media_id)
        media_url = jwt_signed_url(path, jw_key, expires)
        response['jw_player'] = {
            'player_url': signed_player_url, 'media_url': media_url
        }
    return success(response)


@item.route('/<int:item_id>/response', methods=['POST'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def response_process(item_id):
    response = {
        'result': 'success',
    }
    response_json = request.json
    qti_item_obj = Item.query.filter_by(id=item_id).first()
    try:
        item_service = ItemService(qti_item_obj.file_link)
        qti_xml = item_service.get_qti_xml_path()
        response_json['qtiFilename'] = qti_xml
        processing_php = current_app.config['QTI_RSP_PROCESSING_PHP']
        parameter = json.dumps(response_json)
        try:
            result = subprocess.run(['php', processing_php, parameter], stdout=subprocess.PIPE)
            processed = result.stdout.decode("utf-8")
            json.loads(processed)
            response['processed'] = processed
        except Exception as e:
            response['processed'] = "Not implemented."
    except Exception as e:
        print(e)
    response_str = json.dumps(response)
    response = current_app.response_class(response=response_str,
                                          status=200,
                                          mimetype='application/json')
    return response


@item.route('/<int:item_id>/extended', methods=['GET'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def extended_edit(item_id):
    item = Item.query.filter_by(id=item_id).first()
    explanation = ItemExplanation.query.filter_by(item_id=item_id).filter_by(active=True).first()
    form = ItemEditExplanationForm()
    if explanation:
        form.explanation.data = explanation.explanation
        form.link1.data = explanation.links["link1"]
        form.link2.data = explanation.links["link2"]
        form.link3.data = explanation.links["link3"]
        form.link4.data = explanation.links["link4"]
        form.link5.data = explanation.links["link5"]

    return render_template("item/item_edit_explanation.html", form=form, item=item, explanation=explanation)


@item.route('/<int:item_id>/extended', methods=['POST'])
@login_required
@permission_required(Permission.ITEM_MANAGE)
def extended_update(item_id):
    form = ItemEditExplanationForm()
    if form.validate_on_submit():
        img_remove_list = form.chx_remove.data.split(',')

        # Build img_json and link_json data
        images = request.files.getlist("image")
        img_json = []
        link_json = {"link1": form.link1.data,
                     "link2": form.link2.data,
                     "link3": form.link3.data,
                     "link4": form.link4.data,
                     "link5": form.link5.data}

        # Get data from old item explanation
        explanation = ItemExplanation.query.filter_by(item_id=item_id).first()
        if explanation:
            img_json = explanation.images["images"]
            new_img_json = []
            index = 1
            for img in img_json:
                if not (str(index) in img_remove_list):
                    new_img_json.append(img)
                index += 1
            for img_file in images:
                data = img_file.read()
                mimetype = img_file.mimetype
                if data:
                    import base64
                    encodedBytes = base64.b64encode(data)
                    encodedStr = str(encodedBytes, "utf-8")
                    new_json_str = {"mime-type": mimetype, "image": encodedStr}
                    new_img_json.append(new_json_str)
            new_images = {"images": new_img_json}
            explanation.explanation = form.explanation.data
            explanation.links = link_json
            explanation.images = new_images
            flash("Item Explanation has been updated.")
        else:
            for img in images:
                data = img.read()
                mimetype = img.mimetype
                if data:
                    import base64
                    encodedBytes = base64.b64encode(data)
                    encodedStr = str(encodedBytes, "utf-8")
                    json_str = {"mime-type": mimetype, "image": encodedStr}
                    img_json.append(json_str)
            images = {"images": img_json}
            explanation = ItemExplanation(GUID=str(uuid.uuid4()),
                                          item_id=item_id,
                                          version=1,
                                          explanation=form.explanation.data,
                                          images=images,
                                          links=link_json,
                                          modified_by=current_user.id
                                          )
            db.session.add(explanation)
            flash("Item Explanation has been updated.")
        db.session.commit()
        return redirect(url_for('item.manage', grade=explanation.item.grade, subject=explanation.item.subject))
    return redirect(url_for('item.manage', error="Item Edit Explanation- Form validation error"))


def set_choices_child(parent_id):
    choices = Choices.get_codes_child(parent_id)
    return choices


@item.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']


def parse_qti_xml(xml_file, guid, file_name):
    from lxml.etree import XMLSyntaxError

    try:
        loader = ItemLoader(xml_file)
    except XMLSyntaxError:
        return None
    qti_item = loader.build_item()
    parsed = {
        'guid': guid, 'file_link': file_name,
        'qti_item': qti_item
    }
    return parsed


def parse_qti_package_zip(zip_file, file_name):
    from qti.loader.packageloader import PackageLoader

    parsed_items = []

    loader = PackageLoader()
    qti_items = loader.import_qti_package_file(zip_file)
    for qti_item in qti_items:
        parsed = {
            'file_link': file_name,
            'qti_item': qti_item
        }
        try:
            item_service = ItemService(qti_item.get_resource_id())
            qti_xml = item_service.get_qti_xml_path()
            processing_php = current_app.config['QTI_RSP_PROCESSING_PHP']
            parameter = json.dumps({'response': [], 'qtiFilename': qti_xml})
            try:
                result = subprocess.run(['php', processing_php, parameter], stdout=subprocess.PIPE)
                processed = result.stdout.decode("utf-8")
                processed = json.loads(processed)
                parsed['outcome_score'] = processed.get('maxScore')
                parsed['correct_r_value'] = parse_correct_response(processed.get('correctResponses'))
            except Exception as e:
                pass
        except Exception as e:
            print(e)
        my_guid = str(uuid.uuid4())
        parsed['guid'] = my_guid
        parsed_items.append(parsed)
    return parsed_items


def parse_correct_response(correct_response):
    correct_response = correct_response.strip()
    if correct_response != '' and correct_response[0] == '[':
        correct_response = correct_response.replace("'", '"')
        correct_response = json.loads(correct_response)
    return correct_response


def get_identifier():
    random.seed()
    return 'c%s%s' % (str(int(time())), ''.join(str(random.randrange(0, 9)) for i in range(9)))


def generate_qti_package(items):
    """
    Create a qti package using templates
    :param items: List of items
    :return: A path to generate qti package .zip file
    """
    resources = []

    class CHOICE:
        def __init__(self, id, text):
            self.identifier = id
            self.text = text

    class RESOURCE:
        def __init__(self, item):
            self.identifier = get_identifier()
            self.interaction_type = item['interaction_type']
            self.title = item['title']
            self.cardinality = item['cardinality']
            self.baseType = item['baseType']
            self.body = item['body']
            self.prompt = item['prompt']
            self.base = item['base']
            self.patternMask = item['patternMask']
            self.placeholderText = item['placeholderText']
            self.correctResponse = item['correctResponse']
            self.shuffle = item['shuffle']
            self.maxChoices = item['maxChoices']
            self.minChoices = item['minChoices']
            self.orientation = item['orientation']
            self.choices = []

            if item['choices']:
                from string import ascii_uppercase
                choice_identifiers = list(ascii_uppercase)
                for index, value in enumerate(item['choices'].split('\n')):
                    choice = CHOICE(choice_identifiers[index], value)
                    self.choices.append(choice)

    for item in items:
        resources.append(RESOURCE(item))

    manifest_id = get_identifier()
    manifest_xml = render_template('import/imsmanifest.xml', identifier=manifest_id, resources=resources)

    manifest_dir = os.path.join(Config.IMPORT_TEMP_DIR, manifest_id)
    os.mkdir(manifest_dir)
    with open(os.path.join(manifest_dir, 'imsmanifest.xml'), 'w+', encoding='utf-8') as f:
        f.write(manifest_xml)

    for resource in resources:
        qti_xml = render_template('import/%s.xml' % resource.interaction_type, resource=resource)

        qti_dir = os.path.join(manifest_dir, resource.identifier)
        os.mkdir(qti_dir)
        with open(os.path.join(qti_dir, 'qti.xml'), 'w', encoding='utf-8') as f:
            f.write(qti_xml)

    zip_path = os.path.normpath('%s/%s.zip' % (Config.IMPORT_TEMP_DIR, manifest_id))
    shutil.make_archive(zip_path.replace('.zip', ''), 'zip', manifest_dir)

    if not os.path.exists(zip_path):
        zip_path = None
    return zip_path


def get_package_from_excel(excel_file):
    """
    Import an excel file. xlrd can handle both .xls and .xlsx while openpyxl supports .xlsx only
    openpyxl works better so use it for .xlsx
    :param excel_file: A path to an excel file
    :return: A path to a QTI package zip file
    """
    items = []
    if os.path.splitext(excel_file)[-1] == '.xls':
        wb = open_workbook(excel_file, on_demand=True)
        for s in wb.sheets():
            log.info("Sheet: %s" % s.name)
            if s.name == 'Codes':
                continue
            headers = []
            for col in range(s.ncols):
                headers.append(s.cell(0, col).value)
            log.debug("headers: %s" % ",".join(headers))
            for row in range(1, s.nrows):
                a_item = {'identifier': get_identifier()}
                for col in range(s.ncols):
                    column = headers[col]
                    value = s.cell(row, col).value
                    if type(value) == float and value == int(value):
                        value = int(value)
                    a_item[column] = value
                items.append(a_item)
        wb.release_resources()
        return generate_qti_package(items)
    else:
        wb = load_workbook(excel_file)
        for ws in wb:
            log.info("Sheet: %s" % ws.title)
            if ws.title == 'Codes':
                continue
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            log.debug("headers: %s" % ",".join(headers))
            for row in ws.iter_rows(min_row=2):
                empty = all(cell.value is None for cell in row)
                if not empty:
                    a_item = {'identifier': get_identifier()}
                    for cell in row:
                        column = headers[cell.column - 1]
                        value = cell.value
                        a_item[column] = value
                    items.append(a_item)
        return generate_qti_package(items)
